
import asyncio, threading, traceback
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook

from config.settings import settings
from browser.playwright_manager import BrowserManager
from search.search_engine import DuckDuckGoSearch
from company.researcher import Researcher
from excel_writer import IncrementalExcelWriter

class App:
    def __init__(self, root):
        self.root = root
        root.title("Company Research Automation - Production")
        root.geometry("1000x760")

        self.file = tk.StringVar()
        self.col = tk.StringVar()
        self.out = tk.StringVar(
            value=str(Path("output/company_research.xlsx").resolve())
        )
        self.profile = tk.StringVar(
            value=str(Path(settings.profile_dir).resolve())
        )

        # This is now a batch-level browser option only.
        self.keep_browser = tk.BooleanVar(value=settings.keep_browser_open)
        self.resume = tk.BooleanVar(value=True)
        self.linkedin_setup_running = False

        self.build()

    def build(self):
        f = ttk.Frame(self.root, padding=20)
        f.pack(fill="both", expand=True)

        ttk.Label(
            f,
            text="Company Research Automation - Production",
            font=("Segoe UI", 18, "bold")
        ).pack(anchor="w")

        ttk.Label(
            f,
            text=(
                "Excel → DuckDuckGo → Company Website + LinkedIn → "
                "About/Industry/Contact → Incremental Excel"
            )
        ).pack(anchor="w", pady=(2, 15))

        a = ttk.LabelFrame(f, text="Input Excel", padding=12)
        a.pack(fill="x")

        ttk.Entry(a, textvariable=self.file).grid(
            row=0, column=0, sticky="ew", padx=(0, 8)
        )
        ttk.Button(a, text="Browse", command=self.browse).grid(row=0, column=1)

        ttk.Label(a, text="Company Name Column").grid(
            row=1, column=0, sticky="w", pady=(10, 0)
        )

        self.cb = ttk.Combobox(
            a, textvariable=self.col, state="readonly"
        )
        self.cb.grid(
            row=1, column=0, sticky="ew", padx=(160, 8), pady=(10, 0)
        )
        a.columnconfigure(0, weight=1)

        b = ttk.LabelFrame(f, text="Browser", padding=12)
        b.pack(fill="x", pady=12)

        ttk.Label(b, text="Dedicated Chrome profile").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Entry(b, textvariable=self.profile).grid(
            row=0, column=1, sticky="ew", padx=8
        )
        ttk.Button(
            b, text="Browse", command=self.browse_profile
        ).grid(row=0, column=2)
        b.columnconfigure(1, weight=1)

        ttk.Checkbutton(
            b,
            text="Keep main browser open after the entire batch",
            variable=self.keep_browser
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(8, 0))

        ttk.Label(
            b,
            text=(
                "Company website/LinkedIn tabs are always closed after each "
                "company. Only the reusable search tab remains during processing."
            ),
            foreground="#555"
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(7, 0))

        ttk.Checkbutton(
            b,
            text="Resume from existing output Excel",
            variable=self.resume
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(7, 0))

        self.linkedin_button = ttk.Button(
            b,
            text="Setup LinkedIn Login (One-Time)",
            command=self.setup_linkedin_login
        )
        self.linkedin_button.grid(
            row=4, column=0, sticky="w", pady=(10, 0)
        )

        ttk.Label(
            b,
            text=(
                "Optional but recommended: log in to LinkedIn once. "
                "The session is stored in the dedicated browser profile "
                "and reused for future batches."
            ),
            foreground="#555"
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(5, 0))

        c = ttk.LabelFrame(f, text="Output", padding=12)
        c.pack(fill="x")

        ttk.Entry(c, textvariable=self.out).grid(
            row=0, column=0, sticky="ew", padx=(0, 8)
        )
        ttk.Button(c, text="Save As", command=self.saveas).grid(row=0, column=1)
        c.columnconfigure(0, weight=1)

        self.start = ttk.Button(
            f, text="START RESEARCH", command=self.start_job
        )
        self.start.pack(fill="x", pady=14)

        self.progress = ttk.Progressbar(f, mode="determinate")
        self.progress.pack(fill="x")

        self.progress_label = ttk.Label(f, text="0 / 0 companies")
        self.progress_label.pack(anchor="e", pady=(3, 0))

        l = ttk.LabelFrame(f, text="Research Log", padding=8)
        l.pack(fill="both", expand=True, pady=(10, 0))

        self.logbox = tk.Text(
            l, wrap="word", font=("Consolas", 9)
        )
        self.logbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(
            l, command=self.logbox.yview
        )
        scrollbar.pack(side="right", fill="y")
        self.logbox.configure(yscrollcommand=scrollbar.set)

    def log(self, msg):
        print(msg, flush=True)
        self.root.after(0, lambda m=msg: self._log(m))

    def _log(self, msg):
        self.logbox.insert("end", msg + "\n")
        self.logbox.see("end")

    def browse(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xlsm")]
        )
        if not p:
            return

        self.file.set(p)

        wb = load_workbook(
            p, read_only=True, data_only=True
        )
        ws = wb.active

        hs = [
            str(c.value).strip()
            for c in ws[1]
            if c.value is not None
        ]

        wb.close()

        self.cb["values"] = hs
        self.col.set(
            next(
                (
                    x for x in hs
                    if x.lower() in {
                        "company",
                        "company name",
                        "company_name"
                    }
                ),
                hs[0] if hs else ""
            )
        )

    def browse_profile(self):
        p = filedialog.askdirectory()
        if p:
            self.profile.set(p)

    def saveas(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if p:
            self.out.set(p)

    def setup_linkedin_login(self):
        """
        Open LinkedIn in the SAME persistent browser profile used by the
        research batch. The user logs in manually; the application never
        receives or stores the LinkedIn password.
        """
        if self.linkedin_setup_running:
            return

        self.linkedin_setup_running = True
        self.linkedin_button.config(state="disabled")
        self.log("Starting one-time LinkedIn login setup...")

        threading.Thread(
            target=self.linkedin_login_worker,
            daemon=True
        ).start()

    def linkedin_login_worker(self):
        async def run_setup():
            s = type(settings)(
                browser_channel=settings.browser_channel,
                browser_timeout_ms=settings.browser_timeout_ms,
                search_timeout_ms=settings.search_timeout_ms,
                max_search_results=settings.max_search_results,
                max_queries=settings.max_queries,
                max_website_domains=settings.max_website_domains,
                max_internal_pages=settings.max_internal_pages,
                industry_search_results=settings.industry_search_results,
                industry_search_timeout_ms=settings.industry_search_timeout_ms,
                search_assist_open_wait_ms=settings.search_assist_open_wait_ms,
                search_assist_max_wait_ms=settings.search_assist_max_wait_ms,
                search_assist_expanded_wait_ms=settings.search_assist_expanded_wait_ms,
                company_timeout_ms=settings.company_timeout_ms,
                candidate_navigation_attempts=settings.candidate_navigation_attempts,
                linkedin_login_enabled=settings.linkedin_login_enabled,
                keep_browser_open=False,
                use_persistent_profile=True,
                profile_dir=self.profile.get()
            )

            bm = BrowserManager(s)

            try:
                await bm.start()

                p = bm.page
                await p.goto(
                    "https://www.linkedin.com/login",
                    wait_until="domcontentloaded",
                    timeout=s.browser_timeout_ms
                )

                self.log(
                    "LinkedIn login page opened in the dedicated browser profile."
                )
                self.log(
                    "Log in manually. Your password is never read or stored "
                    "by this application."
                )

                done = threading.Event()

                def show_login_message():
                    try:
                        messagebox.showinfo(
                            "LinkedIn Login",
                            (
                                "LinkedIn has been opened in the dedicated "
                                "browser profile.\n\n"
                                "1. Log in to LinkedIn manually.\n"
                                "2. Complete any verification/challenge.\n"
                                "3. Make sure the LinkedIn home/company page "
                                "is accessible.\n"
                                "4. Click OK here to save the browser session."
                            )
                        )
                    finally:
                        done.set()

                self.root.after(0, show_login_message)

                # Wait without blocking Tkinter's UI thread.
                await asyncio.to_thread(done.wait)

                # Give LinkedIn a moment to persist cookies/local storage.
                await asyncio.sleep(1)

                self.log(
                    "LinkedIn session setup completed. "
                    "The persistent browser profile will reuse this login."
                )

            except Exception as exc:
                msg = f"{type(exc).__name__}: {exc}"
                self.log("LinkedIn setup error: " + msg)
                self.root.after(
                    0,
                    lambda m=msg: messagebox.showerror(
                        "LinkedIn Setup Error", m
                    )
                )
            finally:
                await bm.close()

        try:
            asyncio.run(run_setup())
        finally:
            self.root.after(
                0,
                lambda: self.linkedin_button.config(state="normal")
            )
            self.linkedin_setup_running = False

    def start_job(self):
        if not self.file.get() or not self.col.get():
            messagebox.showwarning(
                "Input Required",
                "Select Excel and Company Name column."
            )
            return

        self.start.config(state="disabled")
        self.progress["value"] = 0
        self.logbox.delete("1.0", "end")

        threading.Thread(
            target=self.worker,
            daemon=True
        ).start()

    def worker(self):
        try:
            asyncio.run(self.process())
        except Exception as exc:
            msg = str(exc) or exc.__class__.__name__
            trace = traceback.format_exc()

            print(trace, flush=True)
            self.log("ERROR: " + msg)

            self.root.after(
                0,
                lambda m=msg:
                    messagebox.showerror("Research Error", m)
            )
        finally:
            self.root.after(
                0,
                lambda: self.start.config(state="normal")
            )

    async def process(self):
        wb = load_workbook(
            self.file.get(),
            read_only=True,
            data_only=True
        )
        ws = wb.active

        hs = [
            str(c.value).strip()
            if c.value is not None else ""
            for c in ws[1]
        ]

        idx = hs.index(self.col.get())

        names = []
        seen_input = set()

        for r in ws.iter_rows(min_row=2, values_only=True):
            if idx >= len(r) or r[idx] is None:
                continue

            name = str(r[idx]).strip()
            if not name:
                continue

            key = name.casefold()
            if key in seen_input:
                continue

            seen_input.add(key)
            names.append(name)

        wb.close()

        if not names:
            raise RuntimeError("No company names found.")

        s = type(settings)(
            browser_channel=settings.browser_channel,
            browser_timeout_ms=settings.browser_timeout_ms,
            search_timeout_ms=settings.search_timeout_ms,
            max_search_results=settings.max_search_results,
            max_queries=settings.max_queries,
            max_website_domains=settings.max_website_domains,
            max_internal_pages=settings.max_internal_pages,
            industry_search_results=settings.industry_search_results,
            industry_search_timeout_ms=settings.industry_search_timeout_ms,
            search_assist_open_wait_ms=settings.search_assist_open_wait_ms,
            search_assist_max_wait_ms=settings.search_assist_max_wait_ms,
            search_assist_expanded_wait_ms=settings.search_assist_expanded_wait_ms,
            industry_sector_timeout_ms=settings.industry_sector_timeout_ms,
            company_timeout_ms=settings.company_timeout_ms,
            candidate_navigation_attempts=settings.candidate_navigation_attempts,
            keep_browser_open=self.keep_browser.get(),
            use_persistent_profile=True,
            profile_dir=self.profile.get()
        )

        bm = BrowserManager(s)
        await bm.start()

        writer = IncrementalExcelWriter(
            self.out.get(),
            resume=self.resume.get()
        )

        processed = 0
        skipped = 0
        failed = 0

        try:
            search = DuckDuckGoSearch(
                bm.page,
                s,
                self.log
            )

            research = Researcher(
                bm.context,
                s,
                self.log
            )

            total = len(names)

            self.log(
                f"Batch started: {total} unique companies."
            )
            self.log(
                f"Output checkpoint: {Path(self.out.get()).resolve()}"
            )
            self.log(
                "Each completed company is saved immediately."
            )
            self.log(
                "Company tabs are closed after every company."
            )

            for i, name in enumerate(names, 1):

                if writer.has_company(name):
                    skipped += 1
                    self.log(
                        f"\n[{i}/{total}] {name}"
                    )
                    self.log(
                        "      SKIPPED: already present in output Excel."
                    )
                    self.update_progress(i, total)
                    continue

                self.log("\n" + "=" * 72)
                self.log(f"[{i}/{total}] {name}")
                self.log("=" * 72)

                try:
                    self.log("[1/4] Searching DuckDuckGo...")

                    candidates = await search.search(name)

                    self.log(
                        f"      Found {len(candidates)} unique candidates."
                    )

                    self.log(
                        "[2/4] Matching website → verifying → "
                        "extracting website/LinkedIn..."
                    )

                    r = await asyncio.wait_for(
                        research.run(name, candidates),
                        timeout=s.company_timeout_ms / 1000,
                    )

                    self.log("[3/4] Result:")
                    self.log(
                        f"      Status: {r.status} | "
                        f"Confidence: {r.confidence} | "
                        f"Score: {r.score}"
                    )
                    self.log(
                        f"      Website: {r.website or '-'}"
                    )
                    self.log(
                        f"      LinkedIn: {r.linkedin_url or '-'}"
                    )
                    self.log(
                        f"      Industry: {r.industry or '-'}"
                    )
                    self.log(
                        f"      Email: {r.emails or '-'}"
                    )
                    self.log(
                        f"      Phone: {r.phones or '-'}"
                    )
                    self.log(
                        f"      Address: {r.address or '-'}"
                    )

                    # CRITICAL:
                    # Persist BEFORE starting the next company.
                    writer.append(r)

                    processed += 1

                    self.log("[4/4] Company completed and saved.")
                    self.log(
                        f"      Excel checkpoint saved: "
                        f"{Path(self.out.get()).resolve()}"
                    )

                except Exception as exc:
                    failed += 1

                    self.log(
                        f"      COMPANY ERROR: {type(exc).__name__}: {exc}"
                    )

                    # Create a failed row so a 1000-company batch can continue.
                    from models.company import CompanyResearchResult

                    failed_result = CompanyResearchResult(
                        company_name=name,
                        status="RESEARCH_ERROR",
                        confidence="LOW",
                        score=0,
                        notes=(
                            f"{type(exc).__name__}: {exc}. "
                            "Research failed for this company; batch continued."
                        )
                    )

                    writer.append(failed_result)

                    self.log(
                        "      Error checkpoint saved; continuing with next company."
                    )

                finally:
                    # Always close every company-specific tab before moving on.
                    await bm.close_company_tabs()

                    self.log(
                        "      Company browser tabs closed."
                    )

                self.update_progress(i, total)

            self.log("\n" + "=" * 72)
            self.log("BATCH COMPLETED")
            self.log("=" * 72)
            self.log(f"      Total input companies : {total}")
            self.log(f"      Newly researched      : {processed}")
            self.log(f"      Already completed     : {skipped}")
            self.log(f"      Research errors       : {failed}")
            self.log(
                f"      Output Excel          : "
                f"{Path(self.out.get()).resolve()}"
            )

            self.root.after(
                0,
                lambda: messagebox.showinfo(
                    "Completed",
                    (
                        f"Batch completed.\n\n"
                        f"Total: {total}\n"
                        f"Researched: {processed}\n"
                        f"Skipped: {skipped}\n"
                        f"Errors: {failed}\n\n"
                        f"Output:\n{Path(self.out.get()).resolve()}"
                    )
                )
            )

            if s.keep_browser_open:
                self.log(
                    "Main Chrome browser kept open after batch completion."
                )

        finally:
            writer.close()

            # When production mode is enabled, close the browser at the end.
            if not s.keep_browser_open:
                await bm.close()

    def update_progress(self, i, total):
        value = (i / total) * 100

        self.root.after(
            0,
            lambda v=value, n=i, t=total: (
                self.progress.config(value=v),
                self.progress_label.config(
                    text=f"{n} / {t} companies"
                )
            )
        )

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
