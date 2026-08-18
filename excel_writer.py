
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADERS = [
    "Company Name","About Us","Industry","Sector","Fiscal Revenue",
    "Established Year","No. of Employees","Company LinkedIn URL",
    "Company Website","Email(s)","Phone/Mobile(s)","Address",
    "Contact Us URL","Status","Confidence","Score","Sources","Notes"
]

WIDTHS = [32,80,30,25,25,18,22,45,45,38,32,75,48,22,14,10,90,75]

class IncrementalExcelWriter:
    """
    Production-oriented checkpoint writer.

    The workbook is created once and then saved after every completed company.
    If the application is interrupted, already completed rows remain on disk.
    Existing company names are treated as checkpoints so a rerun can resume.
    """
    def __init__(self, path, resume=True):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.resume = resume
        self.wb = None
        self.ws = None
        self.completed = set()
        self._open()

    def _open(self):
        if self.path.exists() and self.resume:
            self.wb = load_workbook(self.path)
            self.ws = self.wb["Company Research"] if "Company Research" in self.wb.sheetnames else self.wb.active
            if self.ws.max_row == 0:
                self.ws.append(HEADERS)
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    self.completed.add(str(row[0]).strip().casefold())
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Company Research"
            self.ws.append(HEADERS)
            self._style_header()
            self._save()

    def _style_header(self):
        fill = PatternFill("solid", fgColor="1F4E78")
        for c in self.ws[1]:
            c.fill = fill
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        for i, width in enumerate(WIDTHS, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
        self.ws.freeze_panes = "A2"
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1,self.ws.max_row)}"

    def has_company(self, name):
        return str(name).strip().casefold() in self.completed

    def append(self, r):
        row = [
            r.company_name, r.about_us, r.industry, r.sector,
            r.fiscal_revenue, r.established_year, r.employees,
            r.linkedin_url, r.website, r.emails, r.phones, r.address,
            r.contact_page, r.status, r.confidence, r.score,
            r.sources, r.notes
        ]
        self.ws.append(row)
        for c in self.ws[self.ws.max_row]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        self.completed.add(str(r.company_name).strip().casefold())
        self._save()

    def _save(self):
        # Save a complete valid XLSX checkpoint after each company.
        self._style_header()
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{self.ws.max_row}"
        self.wb.save(self.path)

    def close(self):
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None

# Backward-compatible one-shot writer.
def write(path, results):
    writer = IncrementalExcelWriter(path, resume=False)
    try:
        for r in results:
            writer.append(r)
    finally:
        writer.close()
